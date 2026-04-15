"""Generate a sample Excel file to test the WBR bridge agent."""
import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WK14"

headers = [
    "MP", "DS", "Week", "Trending T4W Scan Compliance (W-11 to W-14)",
    "Wk-14 Scan Compliance", "Deep-dive (Wk-13)",
    "Pickup to Stow (<2 days) WK-1", "Pick up to Depart (<7 days) WK-1",
    "DD on RTS (Wk-1)", "Bridge",
    "Performance will improve from which week?", "POC"
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font

sample_data = [
    ["FR", "DFR1", "WK14", "87%", "82%", "Scanner device issues reported by DA",
     "95%", "90%", "5%", "Device malfunction on 12 scanners impacting scan at delivery. IT replacement in progress.",
     "WK16", "Jean Dupont"],
    ["DE", "DDE5", "WK14", "91%", "85%", "New DA onboarding wave with low training completion",
     "92%", "88%", "3%", "30 new DAs onboarded in WK13 with incomplete CRET scan training. Retraining scheduled.",
     "WK15", "Hans Mueller"],
    ["IT", "DIT3", "WK14", "88%", "79%", "Volume spike from Prime Day returns",
     "85%", "82%", "8%", "150% volume spike on returns. Staffing not adjusted. Requesting additional HC for WK15.",
     "WK16", "Marco Rossi"],
    ["ES", "DES2", "WK14", "90%", "84%", "Process non-compliance at pickup",
     "93%", "91%", "4%", "DAs skipping scan at pickup due to time pressure. Station manager reinforcing SOP compliance.",
     "WK15", "Maria Garcia"],
    ["UK", "DUK7", "WK14", "86%", "80%", "App crash during scan workflow",
     "88%", "85%", "6%", "Rabbit app crash affecting scan workflow. Tech ticket raised. Workaround communicated.",
     "WK16", "James Smith"],
    ["FR", "DFR4", "WK14", "89%", "83%", "Backlog of RTS packages",
     "90%", "87%", "7%", "Aging RTS backlog from WK12 still pending. Additional sort capacity deployed.",
     "WK15", "Pierre Martin"],
    ["DE", "DDE9", "WK14", "92%", "86%", "Scanner connectivity issues in rural routes",
     "94%", "90%", "2%", "Poor mobile connectivity on 8 rural routes causing scan upload failures. Offline mode enabled.",
     "WK15", "Anna Schmidt"],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)

# Auto-width
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

wb.save("/home/mousonko/.workspace/wbr-bridge-agent/sample_compliance.xlsx")
print("Sample Excel created: sample_compliance.xlsx")
