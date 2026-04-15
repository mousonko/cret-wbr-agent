"""Parse DS Bridging Scan Compliance Excel and extract flagged site data."""
import openpyxl
from dataclasses import dataclass


@dataclass
class SiteEntry:
    mp: str
    ds: str
    week: str
    trending_scan_compliance: str
    wk14_scan_compliance: str
    deep_dive: str
    pickup_to_stow: str
    pickup_to_depart: str
    dd_on_rts: str
    bridge: str
    improvement_week: str
    poc: str


# Normalized header keywords mapped to SiteEntry fields
# Ordered list: checked top-to-bottom, first match wins.
# Put more specific patterns before general ones.
HEADER_MAP = [
    ("trending", "trending_scan_compliance"),
    ("t4w", "trending_scan_compliance"),
    ("wk-14 scan", "wk14_scan_compliance"),
    ("wk14 scan", "wk14_scan_compliance"),
    ("scan compliance", "wk14_scan_compliance"),
    ("deep-dive", "deep_dive"),
    ("deep dive", "deep_dive"),
    ("pickup to stow", "pickup_to_stow"),
    ("pick up to stow", "pickup_to_stow"),
    ("pickup to depart", "pickup_to_depart"),
    ("pick up to depart", "pickup_to_depart"),
    ("dd on rts", "dd_on_rts"),
    ("bridge", "bridge"),
    ("improve", "improvement_week"),
    ("poc", "poc"),
    ("mp", "mp"),
    ("ds", "ds"),
    ("week", "week"),
]


def _match_header(header_text: str) -> str | None:
    h = str(header_text).lower().strip()
    for keyword, field in HEADER_MAP:
        if keyword in h:
            return field
    return None


def parse_excel(filepath: str, sheet_name: str | None = None) -> list[SiteEntry]:
    """Parse the Excel file and return a list of SiteEntry for flagged sites.

    If sheet_name is None, uses the last sheet (latest week).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        # Pick last sheet as latest week
        ws = wb[wb.sheetnames[-1]]
        sheet_name = ws.title

    print(f"Reading sheet: {sheet_name}")

    # Find header row by scanning first 10 rows for "DS" or "Bridge"
    header_row = None
    col_map = {}
    for row_idx in range(1, 11):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                field = _match_header(str(val))
                if field:
                    col_map[field] = col_idx
        if "ds" in col_map and "bridge" in col_map:
            header_row = row_idx
            break
        col_map.clear()

    if header_row is None:
        raise ValueError(f"Could not find header row in sheet '{sheet_name}'")

    print(f"Header row: {header_row}, mapped columns: {list(col_map.keys())}")

    entries = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ds_val = ws.cell(row=row_idx, column=col_map.get("ds", 1)).value
        if not ds_val:
            continue

        def cell(field):
            col = col_map.get(field)
            if col is None:
                return ""
            v = ws.cell(row=row_idx, column=col).value
            return str(v).strip() if v is not None else ""

        entry = SiteEntry(
            mp=cell("mp"),
            ds=cell("ds"),
            week=cell("week"),
            trending_scan_compliance=cell("trending_scan_compliance"),
            wk14_scan_compliance=cell("wk14_scan_compliance"),
            deep_dive=cell("deep_dive"),
            pickup_to_stow=cell("pickup_to_stow"),
            pickup_to_depart=cell("pickup_to_depart"),
            dd_on_rts=cell("dd_on_rts"),
            bridge=cell("bridge"),
            improvement_week=cell("improvement_week"),
            poc=cell("poc"),
        )
        entries.append(entry)

    print(f"Parsed {len(entries)} site entries")
    return entries
